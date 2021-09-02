#%%
from collections import defaultdict
from datetime import date, datetime, time, timedelta
from datetime import time as time1
from pathlib import Path
from openpyxl.utils import get_column_letter
from dataclasses import dataclass

from functools import partial

import openpyxl
import numpy as np
from pandas import DataFrame
import plotly.graph_objects as go
from plotly.subplots import make_subplots

from vnpy.trader.constant import (Direction, Exchange, Interval, Offset)

@dataclass
class TradeData:
    """
    Trade data contains information of a fill of an order. One order
    can have several trade fills.
    """

    symbol: str
    orderid: str
    tradeid: str
    direction: Direction = None

    offset: Offset = Offset.NONE
    price: float = 0
    volume: float = 0
    datetime: datetime = None


class PnlCaculate:
    """"""

    def __init__(self, real_strategy_name: str):
        """"""
        self.vt_symbol = ""
        self.symbol = ""
        self.exchange = None
        self.start = None
        self.end = None
        self.rate = 0
        self.slippage = 0
        self.size = 1
        self.pricetick = 0
        self.capital = 1_000_000
        self.risk_free: float = 0.02

        self.strategy_class = None
        self.strategy = None
        self.datetime = None

        self.interval = None
        self.days = 0
        self.callback = None
        self.history_data = []

        self.stop_order_count = 0
        self.stop_orders = {}
        self.active_stop_orders = {}

        self.limit_order_count = 0
        self.limit_orders = {}
        self.active_limit_orders = {}

        self.trade_count = 0
        self.trades = {}

        self.logs = []

        self.daily_results = {}
        self.daily_df = None

        self.path = Path(Path.cwd()/"PaperAccount_reord_table.xlsx")
        self.trade_record_dict = {}

        self.strategy_name = real_strategy_name
     
        self.pnl_list = []
        self.total_pnl = 0
        
        self.trade_date = None
        
        self.daily_trades = defaultdict(list)
        self.daily_sum = {}
        
        self.day_trade = False
        self.night_trade = False

        self.daily_pnl = DataFrame()

    def set_parameters(
        self,
        vt_symbol: str,
        interval: Interval,
        start: datetime,
        rate: float,
        slippage: float,
        size: float,
        pricetick: float,
        capital: int = 0,
        end: datetime = None,
        annual_days: int = 240
    ):
        """"""
        
        self.vt_symbol = vt_symbol
        self.rate = rate
        self.slippage = slippage
        self.size = size
        self.start = start

        self.symbol, exchange_str = self.vt_symbol.split(".")
        self.exchange = Exchange(exchange_str)

        self.capital = capital
        self.end = end
        self.annual_days = annual_days

    def get_trade_record(self):
        """"""
        self.trade_record_wb = openpyxl.load_workbook(self.path)

        sheet_names = self.trade_record_wb.sheetnames
        
        if self.strategy_name not in sheet_names:
            print(f"sheet:{self.strategy_name} not in trade record")
            return
        else:
            self.trade_record_sheet = self.trade_record_wb[self.strategy_name]

        total_row = self.trade_record_sheet.max_row
        # total_column = self.trade_record_sheet.max_column

        for i in range(2, total_row + 1):

            trade_datetime = self.trade_record_sheet.cell(i, 8).value
            trade_datetime = trade_datetime.split("+")[0]
            
            weekday = datetime.strptime(trade_datetime, "%Y-%m-%d %H:%M:%S.%f").weekday()

            trade = TradeData(
                    symbol=self.trade_record_sheet.cell(i, 1).value,
                    orderid=self.trade_record_sheet.cell(i, 2).value,
                    tradeid=self.trade_record_sheet.cell(i, 3).value,
                    offset=self.trade_record_sheet.cell(i, 4).value,
                    direction=self.trade_record_sheet.cell(i, 5).value,
                    price=self.trade_record_sheet.cell(i, 6).value,
                    volume=self.trade_record_sheet.cell(i, 7).value,
                    datetime=datetime.strptime(trade_datetime, "%Y-%m-%d %H:%M:%S.%f")
                )

            if trade.datetime.time() < time1(15, 5):
                self.day_trade = True
                self.night_trade = False
            elif trade.datetime.time() >= time1(20, 0):
                self.day_trade = False
                self.night_trade = True

            d = trade.datetime.date()

            if trade.direction == "Direction.LONG":
                trade_money = -(trade.price * trade.volume * 10 + trade.volume * 0.1)
            elif trade.direction == "Direction.SHORT": 
                trade_money = trade.price * trade.volume * 10 - trade.volume * 0.1

            if weekday < 4:
                if not self.trade_date or d == self.trade_date:
                    if self.day_trade:
                        self.daily_trades[d].append(trade_money)
                    elif self.night_trade:
                        self.daily_trades[d+timedelta(days=1)].append(trade_money)
                else:
                    self.daily_trades[d].append(trade_money)
            else:
                if not self.trade_date or d == self.trade_date:
                    if self.day_trade:
                        self.daily_trades[d].append(trade_money)
                    elif self.night_trade:
                        self.daily_trades[d+timedelta(days=3)].append(trade_money)
                else:
                    self.daily_trades[d].append(trade_money)
            
        for key, value in self.daily_trades.items():
            if (len(value) & 1) != 0:
                print("The length of list is odd number")
                print(value, f"日期为{key}")

            self.daily_sum[key] = sum(value)

        self.daily_pnl = DataFrame(self.daily_sum, index=[0]).T

        print(self.daily_pnl)

        self.trade_date = trade.datetime.date()

        print("逐日盯市盈亏计算完成")
            
        print("开始计算策略统计指标")

            # self.daily_pnl["balance"] = self.daily_pnl[""]

            # self.trades[trade.tradeid] = trade

    # def calculate_result(self):
    #     """"""
    #     print("开始计算逐日盯市盈亏")
        
    #     if not self.trades:
    #         print("成交记录为空，无法计算")
    #         return
        
    #     for trade in self.trades.values():
            
    #         d = trade.datetime.date()
            
            

    #         if trade.direction == "Direction.LONG":
    
    #             trade_money = trade.price * trade.volume * 10 + trade.volume * 0.1

    #             self.pnl_list.append(-trade_money)

    #         elif trade.direction == "Direction.SHORT":

    #             trade_money = trade.price * trade.volume * 10 - trade.volume * 0.1

    #             self.pnl_list.append(trade_money)
            
    #         # if trade.direction == "Direction.LONG":
    
    #         #     trade_money = (trade.price + 0.2) * trade.volume * 10 + trade.volume * 0.1

    #         #     self.pnl_list.append(-trade_money)

    #         # elif trade.direction == "Direction.SHORT":

    #         #     trade_money = (trade.price - 0.2) * trade.volume * 10 - trade.volume * 0.1

    #         #     self.pnl_list.append(trade_money)

    #     print(f"self.pnl_list:{self.pnl_list}")

    #     n = len(self.pnl_list)

    #     # self.total_pnl = sum(self.pnl_list)
    #     # print(f"net_profit:{self.total_pnl}")

    #     if (n & 1) == 0:
    #         self.total_pnl = sum(self.pnl_list)
    #         print(f"net_profit:{self.total_pnl}")
    #     else:
    #         self.pnl_list.remove(self.pnl_list[0])
    #         self.total_pnl = sum(self.pnl_list)
    #         print(f"net_profit:{self.total_pnl}")


pnl = PnlCaculate("papertest1")
pnl.get_trade_record()
# pnl.calculate_result()

# pnl = PnlCaculate("papertest2")
# pnl.get_trade_record()
# pnl.calculate_result()

# pnl = PnlCaculate("papertest3")
# pnl.get_trade_record()
# pnl.calculate_result()


class DailyResult:
    """"""

    def __init__(self, date: date, close_price: float):
        """"""
        self.date = date
        self.close_price = close_price
        self.pre_close = 0

        self.trades = []
        self.trade_count = 0

        self.start_pos = 0
        self.end_pos = 0

        self.turnover = 0
        self.commission = 0
        self.slippage = 0

        self.trading_pnl = 0
        self.holding_pnl = 0
        self.total_pnl = 0
        self.net_pnl = 0

    def add_trade(self, trade: TradeData):
        """"""
        self.trades.append(trade)

    def calculate_pnl(
        self,
        pre_close: float,
        start_pos: float,
        size: int,
        rate: float,
        slippage: float,
        inverse: bool
    ):
        """"""
        
        
        # If no pre_close provided on the first day,
        # use value 1 to avoid zero division error
        if pre_close:
            self.pre_close = pre_close
        else:
            self.pre_close = 1

        # Holding pnl is the pnl from holding position at day start
        self.start_pos = start_pos
        self.end_pos = start_pos

        if not inverse:     # For normal contract
            self.holding_pnl = self.start_pos * \
                (self.close_price - self.pre_close) * size
        else:               # For crypto currency inverse contract
            self.holding_pnl = self.start_pos * \
                (1 / self.pre_close - 1 / self.close_price) * size

        # Trading pnl is the pnl from new trade during the day
        self.trade_count = len(self.trades)

        for trade in self.trades:
            if trade.direction == Direction.LONG:
                pos_change = trade.volume
            else:
                pos_change = -trade.volume

            self.end_pos += pos_change

            # For normal contract
            if not inverse:
                turnover = trade.volume * size * trade.price
                self.trading_pnl += pos_change * \
                    (self.close_price - trade.price) * size
                self.slippage += trade.volume * size * slippage
            # For crypto currency inverse contract
            else:
                turnover = trade.volume * size / trade.price
                self.trading_pnl += pos_change * \
                    (1 / trade.price - 1 / self.close_price) * size
                self.slippage += trade.volume * size * slippage / (trade.price ** 2)

            self.turnover += turnover
            self.commission += turnover * rate

        # Net pnl takes account of commission and slippage cost
        self.total_pnl = self.trading_pnl + self.holding_pnl
        self.net_pnl = self.total_pnl - self.commission - self.slippage