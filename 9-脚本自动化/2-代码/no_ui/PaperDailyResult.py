#%%
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from openpyxl.utils import get_column_letter
from dataclasses import dataclass

from functools import partial

import openpyxl
import numpy as np
from pandas import DataFrame
import plotly.graph_objects as go
from plotly.subplots import make_subplots

from vnpy.trader.constant import (Direction, Exchange, Interval, Offset)

@dataclass
class TradeData:
    """
    Trade data contains information of a fill of an order. One order
    can have several trade fills.
    """

    symbol: str
    orderid: str
    tradeid: str
    direction: Direction = None

    offset: Offset = Offset.NONE
    price: float = 0
    volume: float = 0
    datetime: datetime = None


class PnlCaculate:
    """"""

    def __init__(self, real_strategy_name: str):
        """"""
        self.vt_symbol = ""
        self.symbol = ""
        self.exchange = None
        self.start = None
        self.end = None
        self.rate = 0
        self.slippage = 0
        self.size = 1
        self.pricetick = 0
        self.capital = 1_000_000
        self.risk_free: float = 0.02

        self.strategy_class = None
        self.strategy = None
        self.datetime = None

        self.interval = None
        self.days = 0
        self.callback = None
        self.history_data = []

        self.stop_order_count = 0
        self.stop_orders = {}
        self.active_stop_orders = {}

        self.limit_order_count = 0
        self.limit_orders = {}
        self.active_limit_orders = {}

        self.trade_count = 0
        self.trades = {}

        self.logs = []

        self.daily_results = {}
        self.daily_df = None

        self.path = Path(Path.cwd()/"strategies"/"PaperAccount_reord_table.xlsx")
        self.trade_record_dict = {}

        self.strategy_name = real_strategy_name
     
        self.pnl_list = []
        self.total_pnl = 0

    def set_parameters(
        self,
        vt_symbol: str,
        interval: Interval,
        start: datetime,
        rate: float,
        slippage: float,
        size: float,
        pricetick: float,
        capital: int = 0,
        end: datetime = None,
        annual_days: int = 240
    ):
        """"""
        
        self.vt_symbol = vt_symbol
        self.rate = rate
        self.slippage = slippage
        self.size = size
        self.start = start

        self.symbol, exchange_str = self.vt_symbol.split(".")
        self.exchange = Exchange(exchange_str)

        self.capital = capital
        self.end = end
        self.annual_days = annual_days

    def get_trade_record(self):
        """"""
        self.trade_record_wb = openpyxl.load_workbook(self.path)

        sheet_names = self.trade_record_wb.sheetnames
        
        if self.strategy_name not in sheet_names:
            print(f"sheet:{self.strategy_name} not in trade record")
            return
        else:
            self.trade_record_sheet = self.trade_record_wb[self.strategy_name]

        total_row = self.trade_record_sheet.max_row
        # total_column = self.trade_record_sheet.max_column

        for i in range(2, total_row + 1):

            trade_datetime = self.trade_record_sheet.cell(i, 8).value
            trade_datetime = trade_datetime.split("+")[0]

            trade = TradeData(
                    symbol=self.trade_record_sheet.cell(i, 1).value,
                    orderid=self.trade_record_sheet.cell(i, 2).value,
                    tradeid=self.trade_record_sheet.cell(i, 3).value,
                    offset=self.trade_record_sheet.cell(i, 4).value,
                    direction=self.trade_record_sheet.cell(i, 5).value,
                    price=self.trade_record_sheet.cell(i, 6).value,
                    volume=self.trade_record_sheet.cell(i, 7).value,
                    datetime=datetime.strptime(trade_datetime, "%Y-%m-%d %H:%M:%S.%f")
                )

            print(trade.datetime)

            self.trades[trade.tradeid] = trade

    def calculate_result(self):
        """"""
        # Add trade data into daily reuslt.
        for trade in self.trades.values():

            if trade.direction == "Direction.LONG":
    
                trade_money = trade.price * trade.volume * 10 + trade.volume * 0.1

                self.pnl_list.append(-trade_money)

            elif trade.direction == "Direction.SHORT":

                trade_money = trade.price * trade.volume * 10 - trade.volume * 0.1

                self.pnl_list.append(trade_money)
            
            # if trade.direction == "Direction.LONG":
    
            #     trade_money = (trade.price + 0.2) * trade.volume * 10 + trade.volume * 0.1

            #     self.pnl_list.append(-trade_money)

            # elif trade.direction == "Direction.SHORT":

            #     trade_money = (trade.price - 0.2) * trade.volume * 10 - trade.volume * 0.1

            #     self.pnl_list.append(trade_money)

        print(f"self.pnl_list:{self.pnl_list}")
 
        self.total_pnl = sum(self.pnl_list)

        print(f"net_profit:{self.total_pnl}")


pnl = PnlCaculate("papertest1")
pnl.get_trade_record()
pnl.calculate_result()