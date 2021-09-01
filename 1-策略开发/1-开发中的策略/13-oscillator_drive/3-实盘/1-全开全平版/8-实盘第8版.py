#%%
import copy
import openpyxl

from openpyxl.utils import get_column_letter
from pathlib import Path
from typing import Callable
from datetime import time as time1
from datetime import datetime

from vnpy_ctastrategy import CtaTemplate
from vnpy_ctastrategy.base import StopOrder, StopOrderStatus

from vnpy.trader.object import TickData, BarData, OrderData, TradeData
from vnpy.trader.constant import Interval, Exchange, Status
from vnpy.trader.utility import BarGenerator, ArrayManager


#%%
class OscillatorRealTrading(CtaTemplate):
    """"""
    author = "Huang Ning"

    boll_window = 30
    boll_dev = 4
    atr_window = 16
    risk_level = 50
    sl_multiplier = 5.699
    dis_open = 10
    interval = 2

    boll_up = 0
    boll_down = 0
    ultosc = 0
    buy_dis = 0
    short_dis = 0
    atr_value = 0
    long_stop = 0
    short_stop = 0
    intra_trade_high = 0
    intra_trade_low = 0

    parameters = [
        "boll_window",
        "boll_dev",
        "atr_window",
        "risk_level",
        "sl_multiplier",
        "dis_open",
        "interval"
    ]

    variables = [
        "boll_up",
        "boll_down",
        "ultosc",
        "buy_dis",
        "short_dis",
        "atr_value",
        "long_stop",
        "short_stop",
        "intra_trade_high",
        "intra_trade_low"
    ]

    def __init__(self, cta_engine, strategy_name, vt_symbol, setting):
        """"""
        super().__init__(cta_engine, strategy_name, vt_symbol, setting)

        self.bg = XminBarGenerator(self.on_bar, self.interval, self.on_xmin_bar)
        self.am = ArrayManager()

        self.liq_price = 0

        self.trading_size = 0

        self.on_bar_time = time1(0, 0)
        self.day_clearance_time = time1(14, 57)
        self.day_liq_time = time1(14, 59)
        # self.night_clearance_time = time1(22, 57)
        # self.night_liq_time = time1(22, 59)

        self.day_clearance = False
        # self.night_clearance = False

        self.path = Path.cwd()
        self.trade_record_dict = {}

        trade_record_fields = [
            "vt_symbol",
            "orderid",
            "tradeid",
            "offset",
            "direction",
            "price",
            "volume",
            "datetime",
            "strategy-name",
            "strategy"
        ]

        self.trade_record_wb = openpyxl.load_workbook(self.path/"strategies"/"trade_reord_table.xlsx")
        self.trade_record_wb.iso_dates = True  # excel表格可以兼容datetime格式

        sheet_names = self.trade_record_wb.sheetnames
        
        if self.strategy_name not in sheet_names:
            self.trade_record_sheet = self.trade_record_wb.create_sheet(index=0, title=self.strategy_name)
        else:
            self.trade_record_sheet = self.trade_record_wb[self.strategy_name]

        if not self.trade_record_sheet.cell(row=1, column=1).value:
            for i in range(1, len(trade_record_fields)+1):
                column = get_column_letter(i)
                self.trade_record_sheet[column+str(1)] = trade_record_fields[i-1]

        self.trade_record_sheet.freeze_panes = "A2"

        self.trade_record_wb.save(self.path/"strategies"/"trade_reord_table.xlsx")

    def on_init(self):
        """"""
        self.write_log("策略初始化")
        self.load_bar(10)

    def on_start(self):
        """"""
        self.write_log("策略启动")

    def on_stop(self):
        """"""
        self.write_log("策略停止")

    def on_tick(self, tick: TickData):
        """"""
        # 显示策略启动过程中收到的前10个tick
        self.count += 1
        if self.count <= 10:
            self.write_log(tick)

        # 过滤掉非交易时段收到的tick
        if (
            (time1(9, 0) <= tick.datetime.time() <= time1(11, 30))
            or (time1(13, 30) <= tick.datetime.time() <= time1(15, 0))
            or (time1(21, 0) <= tick.datetime.time() <= time1(23, 0))
            ):
            
            self.bg.update_tick(tick)

    def on_bar(self, bar: BarData):
        """"""

        self.liq_price = bar.close_price
        self.on_bar_time = bar.datetime.time()  # 这两行代码一定不能放到self.bg.update_bar(bar)之后

        if datetime.today().isoweekday() <=4:
            self.day_clearance = (self.clearance_time <= self.on_bar_time <= self.liq_time)
        elif datetime.today().isoweekday() == 5:
            self.day_clearance = (
                (self.clearance_time <= self.on_bar_time <= self.liq_time)
                or (time(22, 57) <= self.on_bar_time <= time(22, 59))
                )
        
        self.bg.update_bar(bar)

        self.day_clearance = (self.day_clearance_time <= self.on_bar_time <= self.day_liq_time)
        # self.night_clearance = (self.night_clearance_time <= self.on_bar_time <= self.night_liq_time)

        if self.day_clearance:

            self.write_log(f"clearance time, on_bar_time:{self.on_bar_time}")

            if not self.cta_engine.strategy_orderid_map[self.strategy_name]:

                pos = copy.deepcopy(self.pos)
                self.write_log(f"clearance time, no previous commission, self.pos:{pos}")

                if self.pos > 0:
                    self.sell(self.liq_price - 5, abs(self.pos))
                    self.write_log(f"clearance time, on_bar, sell volume:{pos} {self.on_bar_time}")

                elif self.pos < 0:
                    self.cover(self.liq_price + 5, abs(self.pos))
                    self.write_log(f"clearance time, on_bar, cover volume:{pos} {self.on_bar_time}")

            else:
                orders_buf = copy.deepcopy(self.cta_engine.strategy_orderid_map[self.strategy_name])

                if orders_buf:
                    for vt_orderid in orders_buf:
                        self.cancel_order(vt_orderid)
                        self.write_log(f"clearance time, on_bar, cancel {vt_orderid}")

    def on_xmin_bar(self, bar: BarData):
        """"""

        am = self.am
        am.update_bar(bar)
        if not am.inited:
            return

        self.boll_up, self.boll_down = am.boll(self.boll_window, self.boll_dev)

        self.ultosc = am.ultosc()
        self.buy_dis = 50 + self.dis_open
        self.short_dis = 50 - self.dis_open
        self.atr_value = am.atr(self.atr_window)

        if not self.day_clearance:

            pos = copy.deepcopy(self.pos)

            self.write_log(f"on_xmin_bar, self.pos:{pos}, on_bar_time:{self.on_bar_time}")

            if not self.cta_engine.strategy_orderid_map[self.strategy_name]:
                if self.pos == 0:

                    self.trading_size = max(int(self.risk_level / self.atr_value), 1)
                    self.write_log(f"on_xmin_bar, risk_level:{self.risk_level}, atr_value:{self.atr_value}, trading_size:{self.trading_size}")
                    
                    if self.trading_size > 6:
                        self.trading_size = 6

                    self.intra_trade_high = bar.high_price
                    self.intra_trade_low = bar.low_price

                    if self.ultosc > self.buy_dis:
                        self.buy(self.boll_up, self.trading_size, True)
                        self.write_log(f"on_xmin_bar, buy_svt:{list(self.cta_engine.strategy_orderid_map[self.strategy_name])}, volume:{self.trading_size}")

                    elif self.ultosc < self.short_dis:
                        self.short(self.boll_down, self.trading_size, True)
                        self.write_log(f"on_xmin_bar, short_svt:{list(self.cta_engine.strategy_orderid_map[self.strategy_name])}, volume:{self.trading_size}")

                elif self.pos > 0:
                    self.intra_trade_high = max(self.intra_trade_high, bar.high_price)
                    self.intra_trade_low = bar.low_price

                    self.long_stop = self.intra_trade_high - self.atr_value * self.sl_multiplier

                    self.sell(self.long_stop, abs(self.pos), True)
                    self.write_log(f"on_xmin_bar, sell_svt:{list(self.cta_engine.strategy_orderid_map[self.strategy_name])}, volume:{pos}")

                else:
                    self.intra_trade_high = bar.high_price
                    self.intra_trade_low = min(self.intra_trade_low, bar.low_price)

                    self.short_stop = self.intra_trade_low + self.atr_value * self.sl_multiplier

                    self.cover(self.short_stop, abs(self.pos), True)
                    self.write_log(f"on_xmin_bar, cover_svt:{list(self.cta_engine.strategy_orderid_map[self.strategy_name])}, volume:{pos}")

            else:
                orders_buf = copy.deepcopy(self.cta_engine.strategy_orderid_map[self.strategy_name])

                if orders_buf:
                    for vt_orderid in orders_buf:
                        self.cancel_order(vt_orderid)
                        self.write_log(f"on_xmin_bar, cancel {vt_orderid}")

        self.sync_data()  # 防止出现宕机数据丢失
        self.put_event()

    def on_stop_order(self, stop_order: StopOrder):
        """"""

        on_stop_order_time = datetime.now().time()

        self.write_log(f"on_stop_order, {stop_order.stop_orderid} {stop_order.status} {stop_order.offset} {stop_order.direction}, on_stop_order_time:{on_stop_order_time}")

        if stop_order.status == StopOrderStatus.WAITING:
            return

        if stop_order.status == StopOrderStatus.CANCELLED:
        
            if not self.day_clearance:

                if not self.cta_engine.strategy_orderid_map[self.strategy_name]:

                    if self.pos == 0:
                        if self.ultosc > self.buy_dis:
                            self.buy(self.boll_up, self.trading_size, True)
                            self.write_log(f"on_stop_order, buy_svt:{list(self.cta_engine.strategy_orderid_map[self.strategy_name])}, volume:{self.trading_size}")

                        elif self.ultosc < self.short_dis:
                            self.short(self.boll_down, self.trading_size, True)
                            self.write_log(f"on_stop_order, short_svt:{list(self.cta_engine.strategy_orderid_map[self.strategy_name])}, volume:{self.trading_size}")

                    elif self.pos > 0:
                        pos = copy.deepcopy(self.pos)

                        self.sell(self.long_stop, abs(self.pos), True)
                        self.write_log(f"on_stop_order, sell_svt:{list(self.cta_engine.strategy_orderid_map[self.strategy_name])}, volume:{pos}")

                    else:
                        pos = copy.deepcopy(self.pos)

                        self.cover(self.short_stop, abs(self.pos), True)
                        self.write_log(f"on_stop_order, cover_svt:{list(self.cta_engine.strategy_orderid_map[self.strategy_name])}, volume:{pos}")

            else:
                pos = copy.deepcopy(self.pos)

                if not self.cta_engine.strategy_orderid_map[self.strategy_name]:

                    if self.pos > 0:
                        self.sell(self.liq_price - 5, abs(self.pos))
                        self.write_log(f"clearance time, on_stop_order, sell volume:{pos}, on_bar_time:{self.on_bar_time}")

                    elif self.pos < 0:
                        self.cover(self.liq_price + 5, abs(self.pos))
                        self.write_log(f"clearance time, on_stop_order, cover volume:{pos}, on_bar_time:{self.on_bar_time}")

        self.put_event()

    def on_order(self, order: OrderData):
        """"""

        on_order_time = datetime.now().time()

        self.write_log(f"on_order, {order.orderid} {order.status} {order.offset} {order.direction}, on_order_time:{on_order_time}")

        # ACTIVE_STATUSES = set([Status.SUBMITTING, Status.NOTTRADED, Status.PARTTRADED])
        if order.is_active():
            return

        # not ACTIVE_STATUSES = set([Status.ALLTRADED, Status.CANCELLED, Status.REJECTED])
        if order.status in [Status.CANCELLED, Status.REJECTED]:
        
            if not self.day_clearance:

                if not self.cta_engine.strategy_orderid_map[self.strategy_name]:

                    if self.pos == 0:
                        if self.ultosc > self.buy_dis:
                            self.buy(self.boll_up, self.trading_size, True)
                            self.write_log(f"on_order, buy_svt:{list(self.cta_engine.strategy_orderid_map[self.strategy_name])}, volume:{self.trading_size}")

                        elif self.ultosc < self.short_dis:
                            self.short(self.boll_down, self.trading_size, True)
                            self.write_log(f"on_order, short_svt:{list(self.cta_engine.strategy_orderid_map[self.strategy_name])}, volume:{self.trading_size}")

                    elif self.pos > 0:
                        pos = copy.deepcopy(self.pos)

                        self.sell(self.long_stop, abs(self.pos), True)
                        self.write_log(f"on_order, sell_svt:{list(self.cta_engine.strategy_orderid_map[self.strategy_name])}, volume:{pos}")

                    else:
                        pos = copy.deepcopy(self.pos)

                        self.cover(self.short_stop, abs(self.pos), True)
                        self.write_log(f"on_order, cover_svt:{list(self.cta_engine.strategy_orderid_map[self.strategy_name])}, volume:{pos}")

            else:
                pos = copy.deepcopy(self.pos)

                if not self.cta_engine.strategy_orderid_map[self.strategy_name]:

                    if self.pos > 0:
                        self.sell(self.liq_price - 5, abs(self.pos))
                        self.write_log(f"clearance time, on_order, sell volume:{pos}, on_bar_time:{self.on_bar_time}")

                    elif self.pos < 0:
                        self.cover(self.liq_price + 5, abs(self.pos))
                        self.write_log(f"clearance time, on_order, cover volume:{pos}, on_bar_time:{self.on_bar_time}")

        self.put_event()

    def on_trade(self, trade: TradeData):
        """"""

        subject = f"{self.strategy_name} trade notice, trade_time:{trade.datetime}"
        msg = f"trading record:{trade.vt_symbol}\n{trade.orderid}\n{trade.offset}\n{trade.direction}\n{trade.price}\n{trade.volume}\ntrade_time:{trade.datetime}"
        
        self.write_log(msg)

        self.cta_engine.main_engine.send_email(subject, msg)

        self.trade_record_dict = {
            "vt_symbol": str(trade.vt_symbol),
            "orderid": str(trade.orderid),
            "tradeid": str(trade.tradeid),
            "offset": str(trade.offset),
            "direction": str(trade.direction),
            "price": str(trade.price),
            "volume": str(trade.volume),
            "datetime": str(trade.datetime),
            "strategy_name": str(self.strategy_name),
            "strategy": str(self.cta_engine.strategies[self.strategy_name]) 
        }

        # 如果没有明确指定sheet，交易记录会默认保存至第一个sheet
        self.trade_record_sheet = self.trade_record_wb[self.strategy_name]
        
        self.trade_record_sheet.insert_rows(2)

        for i in range(1, self.trade_record_sheet.max_column+1):
            column = get_column_letter(i)
            self.trade_record_sheet[column+str(2)] = list(self.trade_record_dict.values())[i-1]

        self.trade_record_wb.save(self.path/"strategies"/"trade_reord_table.xlsx")

        self.write_log("Trade Record Is Saved")

        self.put_event()


class XminBarGenerator(BarGenerator):
    def __init__(
        self,
        on_bar: Callable,
        window: int = 0,
        on_window_bar: Callable = None,
        interval: Interval = Interval.MINUTE
    ):
        super().__init__(on_bar, window, on_window_bar, interval)

    def update_bar(self, bar: BarData) ->None:
        """
        Update 1 minute bar into generator
        """
        # If not inited, creaate window bar object
        if not self.window_bar:
            # Generate timestamp for bar data
            if self.interval == Interval.MINUTE:
                dt = bar.datetime.replace(second=0, microsecond=0)
            else:
                dt = bar.datetime.replace(minute=0, second=0, microsecond=0)

            self.window_bar = BarData(
                symbol=bar.symbol,
                exchange=bar.exchange,
                datetime=dt,
                gateway_name=bar.gateway_name,
                open_price=bar.open_price,
                high_price=bar.high_price,
                low_price=bar.low_price
            )
        # Otherwise, update high/low price into window bar
        else:
            self.window_bar.high_price = max(
                self.window_bar.high_price, bar.high_price)
            self.window_bar.low_price = min(
                self.window_bar.low_price, bar.low_price)

        # Update close price/volume into window bar
        self.window_bar.close_price = bar.close_price
        self.window_bar.volume += int(bar.volume)
        self.window_bar.open_interest = bar.open_interest

        # Check if window bar completed
        finished = False

        if self.interval == Interval.MINUTE:
            # x-minute bar
            # if not (bar.datetime.minute + 1) % self.window:
            #     finished = True

            self.interval_count += 1

            if not self.interval_count % self.window:
                finished = True
                self.interval_count = 0

            elif bar.datetime.time() in [time1(10, 14), time1(11, 29), time1(14, 59), time1(22, 59)]:
                if bar.exchange in [Exchange.SHFE, Exchange.DCE, Exchange.CZCE]:
                    finished = True
                    self.interval_count = 0

        elif self.interval == Interval.HOUR:
            if self.last_bar:
                new_hour = bar.datetime.hour != self.last_bar.datetime.hour
                last_minute = bar.datetime.minute == 59
                not_first = self.window_bar.datetime != bar.datetime

                # To filter duplicate hour bar finished condition
                if (new_hour or last_minute) and not_first:
                    # 1-hour bar
                    if self.window == 1:
                        finished = True
                    # x-hour bar
                    else:
                        self.interval_count += 1

                        if not self.interval_count % self.window:
                            finished = True
                            self.interval_count = 0

        if finished:
            self.on_window_bar(self.window_bar)
            self.window_bar = None

        # Cache last bar object
        self.last_bar = bar