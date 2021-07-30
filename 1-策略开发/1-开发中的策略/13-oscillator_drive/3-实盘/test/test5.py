#%%
import openpyxl
from openpyxl.styles import Font

wb = openpyxl.load_workbook("D:/CTA/1-策略开发/1-开发中的策略/14-oscillator_drive/2-实盘/censuspopdata.xlsx")

sheet = wb.active

sheet["A1"] = "Tall row"
sheet["B2"] = "Wide column"

sheet.cell(1, 1).value = "黄柠"



italic24Font = Font(size=24, italic=True)

price_updates = {
    "Garlic": 3.07,
    "Celery": 1.19,
    "Lemon": 1.27}

for row_number in range(2, sheet.max_row):
    produce_name = sheet.cell(row=row_number, column=1).value
    if produce_name in price_updates:
        sheet.cell(row=row_number, column=2).value = price_updates[produce_name]
        sheet.cell(row=row_number, column=2).font = italic24Font

sheet["A"+str(sheet.max_row+1)] = "Total"
sheet["D"+str(sheet.max_row-1)] = "=SUM(D1: D23758)" 

wb.save("D:/CTA/1-策略开发/1-开发中的策略/14-oscillator_drive/2-实盘/updatedProduceSales.xlsx")