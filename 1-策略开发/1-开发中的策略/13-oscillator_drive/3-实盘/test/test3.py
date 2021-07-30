#%%
import openpyxl
import pprint
from openpyxl.utils import get_column_letter, column_index_from_string
import datetime

print("opening workbook...")
wb = openpyxl.load_workbook("D:/CTA/1-策略开发/1-开发中的策略/14-oscillator_drive/2-实盘/censuspopdata.xlsx")

sheet = wb["Population by Census Tract"]

countyData = {}

print("reading rows...")

for i in range(1, 4):
    print(i)

for row in range(2, sheet.max_row +1):
    state = sheet["B"+str(row)].value
    county = sheet["C"+str(row)].value
    pop = sheet["D"+str(row)].value
    countyData.setdefault(state, {})
    countyData[state].setdefault(county, {"tracts": 0, "pop": 0})
    countyData[state][county]["tracts"] += 1
    countyData[state][county]["pop"] += int(pop)

print("Writing results...")

resultFile = open("D:/CTA/1-策略开发/1-开发中的策略/14-oscillator_drive/2-实盘/census2010.py", "w")
resultFile.write("allData=" + pprint.pformat(countyData))
resultFile.close()
print("Done.")
    

