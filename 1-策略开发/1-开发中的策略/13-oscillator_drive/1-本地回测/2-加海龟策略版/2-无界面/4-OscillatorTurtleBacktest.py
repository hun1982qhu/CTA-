#%%
import copy
import time
import datetime
import openpyxl

from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import time as time1
from datetime import datetime as datetime1
from typing import Callable

from vnpy_ctastrategy import CtaTemplate
from vnpy_ctastrategy.base import StopOrder, StopOrderStatus
from vnpy_ctastrategy.backtesting import BacktestingEngine, BacktestingMode, OptimizationSetting

from vnpy.trader.object import TickData, BarData, OrderData, TradeData, Direction
from vnpy.trader.constant import Interval, Exchange, Offset, Status
from vnpy.trader.utility import BarGenerator, ArrayManager


class XminBarGenerator(BarGenerator):
    def __init__(
        self,
        on_bar: Callable,
        window: int = 0,
        on_window_bar: Callable = None,
        interval: Interval = Interval.MINUTE
    ):
        super().__init__(on_bar, window, on_window_bar, interval)

    def update_bar_minute_window(self, bar: BarData) -> None:
        """"""
        # If not inited, create window bar object
        if not self.window_bar:
            dt = bar.datetime.replace(second=0, microsecond=0)
            self.window_bar = BarData(
                symbol=bar.symbol,
                exchange=bar.exchange,
                datetime=dt,
                gateway_name=bar.gateway_name,
                open_price=bar.open_price,
                high_price=bar.high_price,
                low_price=bar.low_price
            )
        # Otherwise, update high/low price into window bar
        else:
            self.window_bar.high_price = max(
                self.window_bar.high_price,
                bar.high_price
            )
            self.window_bar.low_price = min(
                self.window_bar.low_price,
                bar.low_price
            )

        # Update close price/volume into window bar
        self.window_bar.close_price = bar.close_price
        self.window_bar.volume += int(bar.volume)
        self.window_bar.open_interest = bar.open_interest

        # Check if window bar completed
        # if not (bar.datetime.minute + 1) % self.window:
        #     self.on_window_bar(self.window_bar)
        #     self.window_bar = None

        finished = False

        self.interval_count += 1

        if not self.interval_count % self.window:
            finished = True
            self.interval_count = 0
            
        elif bar.datetime.time() in [time1(10, 14), time1(11, 29), time1(14, 59), time1(22, 59)]:

            if bar.exchange in [Exchange.SHFE, Exchange.DCE, Exchange.CZCE]:
                finished = True
                self.interval_count = 0

        if finished:
            self.on_window_bar(self.window_bar)
            self.window_bar = None

        # Cache last bar object
        self.last_bar = bar


class OscillatorTurtleBacktest(CtaTemplate):
    """"""
    author = "Huang Ning"

    boll_window = 57
    boll_dev = 4
    atr_window = 4
    risk_level = 50
    sl_multiplier = 4.8999999999999995
    dis_open = 2
    interval = 10
    fixed_size = 2

    boll_up = 0
    boll_down = 0
    ultosc = 0
    buy_dis = 0
    short_dis = 0
    atr_value = 0
    long_stop = 0
    short_stop = 0
    intra_trade_high = 0
    intra_trade_low = 0
    long_entry = 0
    short_entry = 0
    trade_long_stop = 0
    trade_short_stop = 0

    parameters = [
        "boll_window",
        "boll_dev",
        "atr_window",
        "risk_level",
        "sl_multiplier",
        "dis_open",
        "interval",
        "fixed_size"
    ]

    variables = [
        "boll_up",
        "boll_down",
        "ultosc",
        "buy_dis",
        "short_dis",
        "atr_value",
        "long_stop",
        "short_stop",
        "intra_trade_high",
        "intra_trade_low",
        "long_entry",
        "short_entry",
        "trade_long_stop",
        "trade_short_stop"
    ]

    def __init__(self, cta_engine, strategy_name, vt_symbol, setting):
        """"""
        super().__init__(cta_engine, strategy_name, vt_symbol, setting)

        self.bg = XminBarGenerator(self.on_bar, self.interval, self.on_xmin_bar)
        self.am = ArrayManager()

        self.liq_price = 0
        self.trading_size = 0

        self.on_bar_time = time1(0, 0)
        self.clearance_time = time1(14, 57)  # 清仓开始时间
        self.liq_time = time1(14, 59)  # 交易所结算开始时间

        self.day_clearance = False

        self.buy_svt_orderids = []
        self.sell_svt_orderids = []
        self.short_svt_orderids = []
        self.cover_svt_orderids = []

        self.sell_lvt_orderids = []
        self.cover_lvt_orderids = []

        self.extreme_date_list = []

        # 2021年5月10日有极端行情，收益异常高，不具有普遍参考价值，因此在回测中跳过这一天
        extreme_date_list = ["2021-05-10", "2021-05-19", "2021-05-20"]
        for date in extreme_date_list:
            extreme_date = time.strptime(date, "%Y-%m-%d")
            year, month, day = extreme_date[:3]
            extreme_date = datetime.date(year, month, day)
            self.extreme_date_list.append(extreme_date)

    def on_init(self):
        """"""
        self.write_log("策略初始化")
        self.load_bar(10)

    def on_start(self):
        """"""
        self.write_log("策略启动")

    def on_stop(self):
        """"""
        self.write_log("策略停止")

    def on_tick(self, tick: TickData):
        """"""
        self.bg.update_tick(tick)

    def on_bar(self, bar: BarData):
        """"""

        on_bar_date = bar.datetime.date()

        self.liq_price = bar.close_price
        self.on_bar_time = bar.datetime.time()

        self.day_clearance = (self.clearance_time <= self.on_bar_time <= self.liq_time)

        if on_bar_date not in self.extreme_date_list:

            self.bg.update_bar(bar)

            if self.day_clearance:

                if not self.buy_svt_orderids and not self.short_svt_orderids \
                    and not self.sell_svt_orderids and not self.cover_svt_orderids \
                        and not self.sell_lvt_orderids and not self.cover_lvt_orderids:

                        if self.pos > 0:
                            self.sell_lvt_orderids = self.sell(self.liq_price - 5, abs(self.pos))

                        elif self.pos < 0:
                            self.cover_lvt_orderids = self.cover(self.liq_price + 5, abs(self.pos))

                else:

                    for buf_orderids in [
                        self.buy_svt_orderids,
                        self.short_svt_orderids,
                        self.sell_svt_orderids,
                        self.cover_svt_orderids,
                        self.sell_lvt_orderids,
                        self.cover_lvt_orderids
                        ]:
                        
                        if buf_orderids:                    
                            for vt_orderid in buf_orderids:
                                self.cancel_order(vt_orderid)

        # else:
        #     print(f"bar.datetime:{bar.datetime}")
        #     print(f"on_bar_time:{self.on_bar_time}")
        #     print(f"on_bar_date:{on_bar_date}")
        #     print(f"extreme_date:{self.extreme_date}")

    def on_xmin_bar(self, bar: BarData):
        """"""

        am = self.am
        am.update_bar(bar)
        if not am.inited:
            return

        self.boll_up, self.boll_down = am.boll(self.boll_window, self.boll_dev)

        self.ultosc = am.ultosc()
        self.buy_dis = 50 + self.dis_open
        self.short_dis = 50 - self.dis_open
        self.atr_value = am.atr(self.atr_window)

        if not self.day_clearance:

            if self.pos == 0:

                # self.trading_size = max(int(self.risk_level / self.atr_value), 1)
                # self.write_log(f"on_xmin_bar, risk_level:{self.risk_level}, atr_value:{self.atr_value}, trading_size:{self.trading_size}")
                
                # if self.trading_size > 6:
                #     self.trading_size = 6

                self.intra_trade_high = bar.high_price
                self.intra_trade_low = bar.low_price

                if not self.buy_svt_orderids and not self.short_svt_orderids:
                    
                    if self.ultosc > self.buy_dis:
                        self.send_buy_orders(self.boll_up)

                    elif self.ultosc < self.short_dis:
                        self.send_short_orders(self.boll_down)

                else:

                    for buf_orderids in [
                        self.buy_svt_orderids, 
                        self.short_svt_orderids
                        ]:

                        if buf_orderids:
                            for vt_orderid in buf_orderids:
                                self.cancel_order(vt_orderid)
                                # self.write_log(f"on_xmin_bar, cancel {vt_orderid}")

            elif self.pos > 0:
                
                self.intra_trade_high = max(self.intra_trade_high, bar.high_price)
                self.intra_trade_low = bar.low_price

                self.long_stop = self.intra_trade_high - self.atr_value * self.sl_multiplier
                self.long_stop = max(self.long_stop, self.trade_long_stop, self.boll_down)
                
                if not self.buy_svt_orderids and not self.sell_svt_orderids:

                    self.send_buy_orders(self.boll_up)
                   
                    self.sell_svt_orderids = self.sell(self.long_stop, abs(self.pos), True)

                else:                            

                    for buf_orderids in [
                        self.buy_svt_orderids, 
                        self.sell_svt_orderids
                        ]:

                        if buf_orderids:

                            for vt_orderid in buf_orderids:
                                self.cancel_order(vt_orderid)
                                # self.write_log(f"on_xmin_bar, cancel {vt_orderid}")

            else:

                self.intra_trade_high = bar.high_price
                self.intra_trade_low = min(self.intra_trade_low, bar.low_price)

                self.short_stop = self.intra_trade_low + self.atr_value * self.sl_multiplier
                self.short_stop = min(self.short_stop, self.trade_short_stop, self.boll_up)

                if not self.short_svt_orderids and not self.cover_svt_orderids:

                    self.send_short_orders(self.boll_down)

                    self.cover_svt_orderids = self.cover(self.short_stop, abs(self.pos), True)

                else:

                    for buf_orderids in [
                        self.short_svt_orderids, 
                        self.cover_svt_orderids
                        ]:

                        if buf_orderids:
                            for vt_orderid in buf_orderids:
                                self.cancel_order(vt_orderid)
                                self.write_log(f"on_xmin_bar, cancel {vt_orderid}")

        self.sync_data()  # 防止出现宕机数据丢失
        self.put_event()

    def on_stop_order(self, stop_order: StopOrder):
        """"""

        # self.write_log(f"on_stop_order, {stop_order.stop_orderid} {stop_order.status} {stop_order.offset} {stop_order.direction}, on_stop_order_time:{stop_order.datetime}")
        
        if stop_order.status == StopOrderStatus.WAITING:
            return

        for buf_orderids in [
            self.buy_svt_orderids,
            self.sell_svt_orderids,
            self.short_svt_orderids,
            self.cover_svt_orderids]:
            
            if stop_order.stop_orderid in buf_orderids:
                buf_orderids.remove(stop_order.stop_orderid)   

        if stop_order.status == StopOrderStatus.CANCELLED:

            if not self.day_clearance:

                if self.pos == 0:

                    if not self.buy_svt_orderids and not self.short_svt_orderids:

                        if self.ultosc > self.buy_dis:
                            self.send_buy_orders(self.boll_up)
                            # self.write_log(f"on_stop_order, buy_svt:{self.buy_svt_orderids}, volume:{self.trading_size}")

                        elif self.ultosc < self.short_dis:
                            self.send_short_orders(self.boll_down)
                            # self.write_log(f"on_stop_order, short_svt:{self.short_svt_orderids}, volume:{self.trading_size}")

                elif self.pos > 0:

                    if not self.buy_svt_orderids and not self.sell_svt_orderids:

                            self.send_buy_orders(self.boll_up)

                            self.sell_svt_orderids = self.sell(self.long_stop, abs(self.pos), True)
                            # self.write_log(f"on_xmin_bar, sell_svt:{self.sell_svt_orderids}, volume:{pos}")

                else:

                    if not self.short_svt_orderids and not self.cover_svt_orderids:

                        self.send_short_orders(self.boll_down)
  
                        self.cover_svt_orderids = self.cover(self.short_stop, abs(self.pos), True)
                        # self.write_log(f"on_stop_order, cover_svt:{self.cover_svt_orderids}, volume:{pos}")

            else:

                if not self.buy_svt_orderids and not self.short_svt_orderids \
                    and not self.sell_svt_orderids and not self.cover_svt_orderids \
                        and not self.sell_lvt_orderids and not self.cover_lvt_orderids:
                        
                        if self.pos > 0:
                            self.sell_lvt_orderids = self.sell(self.liq_price - 5, abs(self.pos))
                            # self.write_log(f"clearance time, on_stop_order, sell volume:{pos}, on_bar_time:{self.on_bar_time}")

                        elif self.pos < 0:
                            self.cover_lvt_orderids = self.cover(self.liq_price + 5, abs(self.pos))
                            # self.write_log(f"clearance time, on_stop_order, cover volume:{pos}, on_bar_time:{self.on_bar_time}")
           
        self.put_event()

    def on_order(self, order: OrderData):
        """"""

        # msg = f"on_order, {order.orderid} {order.status} {order.offset} {order.direction}, on_order_time:{order.datetime}"

        # self.write_log(msg)

        if order.status in [Status.SUBMITTING, Status.NOTTRADED, Status.PARTTRADED]:
            return

        for buf_orderids in [
            self.sell_lvt_orderids, 
            self.cover_lvt_orderids
            ]:

            if order.orderid in buf_orderids:
                # self.write_log(f"on_order, {order.orderid} is removed from {buf_orderids}")
                buf_orderids.remove(order.orderid)
    
        # not ACTIVE_STATUSES = set([Status.ALLTRADED, Status.CANCELLED, Status.REJECTED])
        if order.status in [Status.CANCELLED, Status.REJECTED]:
        
            if self.day_clearance:

                # pos = copy.deepcopy(self.pos)

                if not self.buy_svt_orderids and not self.short_svt_orderids \
                    and not self.sell_svt_orderids and not self.cover_svt_orderids \
                        and not self.sell_lvt_orderids and not self.cover_lvt_orderids:
                            
                        if self.pos > 0:
                            self.sell_lvt_orderids = self.sell(self.liq_price - 5, abs(self.pos))
                            # self.write_log(f"clearance time, on_order, sell volume:{pos}, on_bar_time:{self.on_bar_time}")

                        elif self.pos < 0:
                            self.cover_lvt_orderids = self.cover(self.liq_price + 5, abs(self.pos))
                            # self.write_log(f"clearance time, on_order, cover volume:{pos}, on_bar_time:{self.on_bar_time}")

        self.put_event()

    def on_trade(self, trade: TradeData):
        """"""

        if trade.direction == Direction.LONG:
            self.long_entry = trade.price
            self.trade_long_stop = self.long_entry - 2 * self.atr_value
        else:
            self.short_entry = trade.price
            self.trade_short_stop = self.short_entry + 2 * self.atr_value

        # subject = f"{self.strategy_name} trade notice, trade_time:{trade.datetime}"
        # msg = f"trading record:{trade.vt_symbol}\n{trade.orderid}\n{trade.offset}\n{trade.direction}\n{trade.price}\n{trade.volume}\ntrade_time:{trade.datetime}"
        
        # self.write_log(msg)

        # self.cta_engine.main_engine.send_email(subject, msg)

        # self.trade_record_dict = {
        #     "vt_symbol": trade.vt_symbol,
        #     "orderid": trade.orderid,
        #     "tradeid": trade.tradeid,
        #     "offset": str(trade.offset),
        #     "direction": str(trade.direction),
        #     "price": trade.price,
        #     "volume": trade.volume,
        #     "datetime": str(trade.datetime),
        #     "strategy": self.strategy_name
        # }

        # self.trade_record_wb = openpyxl.load_workbook(self.path/"strategies"/"PaperAccount_reord_table.xlsx")
        # self.trade_record_sheet = self.trade_record_wb[self.strategy_name]

        # self.trade_record_sheet.insert_rows(2)

        # for i in range(1, self.trade_record_sheet.max_column+1):
        #     column = get_column_letter(i)
        #     self.trade_record_sheet[column+str(2)] = list(self.trade_record_dict.values())[i-1]

        # self.trade_record_wb.save(self.path/"strategies"/"PaperAccount_reord_table.xlsx")

        # self.write_log(f"{self.strategy_name} Trade Record Is Saved")

        # self.put_event()

    def send_buy_orders(self, price):
        """"""
        t = self.pos / self.fixed_size

        if t < 1:
            self.buy_svt_orderids = self.buy(price, self.fixed_size, True)

        if t < 2:
            self.buy_svt_orderids = self.buy(price + self.atr_value * 0.5, self.fixed_size, True)

        if t < 3:
            self.buy_svt_orderids = self.buy(price + self.atr_value, self.fixed_size, True)

        if t < 3.5:
            self.buy_svt_orderids = self.buy(price + self.atr_value * 1.5, self.fixed_size, True)

        # if t < 5:
        #     self.buy_svt_orderids = self.buy(price + self.atr_value * 2, self.fixed_size, True)

        # if t < 6:
        #     self.buy_svt_orderids = self.buy(price + self.atr_value * 2.5, self.fixed_size, True)

        # if t < 7:
        #     self.buy_svt_orderids = self.buy(price + self.atr_value * 3, self.fixed_size, True)

    def send_short_orders(self, price):
        """"""
        t = self.pos / self.fixed_size

        if t > -1:
            self.short_svt_orderids = self.short(price, 3 * self.fixed_size, True)

        if t > -2:
            self.short_svt_orderids = self.short(price - self.atr_value * 0.5, 2 * self.fixed_size, True)

        if t > -3:
            self.short_svt_orderids = self.short(price - self.atr_value, self.fixed_size, True)

        if t > -3.5:
            self.short_svt_orderids = self.short(price - self.atr_value * 1.5, self.fixed_size, True)

        # if t > -5:
        #     self.short_svt_orderids = self.short(price - self.atr_value * 2, self.fixed_size, True)

        # if t > -6:
        #     self.short_svt_orderids = self.short(price - self.atr_value * 2.5, self.fixed_size, True)

        # if t > -7:
        #     self.short_svt_orderids = self.short(price - self.atr_value * 3, self.fixed_size, True)


#%%
start1 = time.time()
engine = BacktestingEngine()
engine.set_parameters(
    vt_symbol="rb888.SHFE",
    interval="1m",
    start=datetime1(2021, 1, 1),
    end=datetime1(2021, 4, 30),
    rate=0.0001,
    slippage=0.2,
    size=10,
    pricetick=1,
    capital=50000,
    mode=BacktestingMode.BAR
)
engine.add_strategy(OscillatorTurtleBacktest, {})
#%%
start2 = time.time()
engine.load_data()
end2 = time.time()
print(f"加载数据所需时长: {(end2-start2)} Seconds")
#%%
engine.run_backtesting()
#%%
engine.calculate_result()
engine.calculate_statistics()
# 待测试的代码
end1 = time.time()
print(f"单次回测运行时长: {(end1-start1)} Seconds")
engine.show_chart()
#%%
# setting = OptimizationSetting()
# setting.set_target("end_balance")
# setting.add_parameter("bar_window_length", 1, 20, 1)
# setting.add_parameter("cci_window", 3, 10, 1)
# setting.add_parameter("fixed_size", 1, 1, 1)
# setting.add_parameter("sell_multipliaer", 0.80, 0.99, 0.01)
# setting.add_parameter("cover_multiplier", 1.01, 1.20, 0.01)
# setting.add_parameter("pricetick_multiplier", 1, 5, 1)
#%%
# engine.run_optimization(setting, output=True)