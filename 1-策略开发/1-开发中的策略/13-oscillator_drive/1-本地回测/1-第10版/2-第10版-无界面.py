#%%
import datetime
import time
import openpyxl

from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import time as time1
from typing import Callable

from vnpy_ctastrategy import CtaTemplate
from vnpy_ctastrategy.base import StopOrder, StopOrderStatus
from vnpy_ctastrategy.backtesting import BacktestingEngine, BacktestingMode, OptimizationSetting

from vnpy.trader.object import TickData, BarData, OrderData, TradeData
from vnpy.trader.constant import Interval, Exchange, Status
from vnpy.trader.utility import BarGenerator, ArrayManager


#%%
#%%
class OscillatorHNBacktest(CtaTemplate):
    """"""
    author = "Huang Ning"

    boll_window = 3
    boll_dev = 10
    atr_window = 12
    risk_level = 50
    sl_multiplier = 6.29999
    dis_open = 4
    interval = 2

    boll_up = 0
    boll_down = 0
    ultosc = 0
    buy_dis = 0
    short_dis = 0
    atr_value = 0
    long_stop = 0
    short_stop = 0
    intra_trade_high = 0
    intra_trade_low = 0

    parameters = [
        "boll_window",
        "boll_dev",
        "atr_window",
        "risk_level",
        "sl_multiplier",
        "dis_open",
        "interval"
    ]

    variables = [
        "boll_up",
        "boll_down",
        "ultosc",
        "buy_dis",
        "short_dis",
        "atr_value",
        "long_stop",
        "short_stop",
        "intra_trade_high",
        "intra_trade_low"
    ]

    def __init__(self, cta_engine, strategy_name, vt_symbol, setting):
        """"""
        super().__init__(cta_engine, strategy_name, vt_symbol, setting)

        self.bg = XminBarGenerator(self.on_bar, self.interval, self.on_xmin_bar)
        self.am = ArrayManager()

        self.liq_price = 0
        self.trading_size = 0

        self.on_bar_time = time1(0, 0)
        self.clearance_time = time1(14, 57)  # 清仓开始时间
        self.liq_time = time1(14, 59)  # 交易所结算开始时间

        self.day_clearance = False

        self.buy_svt_orderids = []
        self.sell_svt_orderids = []
        self.short_svt_orderids = []
        self.cover_svt_orderids = []

        self.sell_lvt_orderids = []
        self.cover_lvt_orderids = []

        # self.path = Path.cwd()
        # self.trade_record_dict = {}

        # trade_record_fields = [
        #     "vt_symbol",
        #     "orderid",
        #     "tradeid",
        #     "offset",
        #     "direction",
        #     "price",
        #     "volume",
        #     "datetime",
        #     "strategy"
        # ]

        # self.trade_record_wb = openpyxl.load_workbook(self.path/"trade_reord_table.xlsx")
        # self.trade_record_wb.iso_dates = True

        # sheet_names = self.trade_record_wb.sheetnames
        
        # if self.strategy_name not in sheet_names:
        #     self.trade_record_sheet = self.trade_record_wb.create_sheet(index=0, title=self.strategy_name)
        # else:
        #     self.trade_record_sheet = self.trade_record_wb[self.strategy_name]

        # if not self.trade_record_sheet.cell(row=1, column=1).value:
        #     for i in range(1, len(trade_record_fields)+1):
        #         column = get_column_letter(i)
        #         self.trade_record_sheet[column+str(1)] = trade_record_fields[i-1]

        # self.trade_record_sheet.freeze_panes = "A2"

    def on_init(self):
        """"""
        self.write_log("策略初始化")
        self.load_bar(10)

    def on_start(self):
        """"""
        self.write_log("策略启动")

    def on_stop(self):
        """"""
        self.write_log("策略停止")

    def on_tick(self, tick: TickData):
        """"""
        self.bg.update_tick(tick)

    def on_bar(self, bar: BarData):
        """"""

        # if bar.datetime.time() > time1(9, 0):
        #     print(f"bar.datetime:{bar.datetime}")
        #     print(f"bar.datetime.time:{bar.datetime.time()}")

        self.liq_price = bar.close_price
        self.on_bar_time = bar.datetime.time()
        self.on_bar_date = bar.datetime.date()

        # 2021年5月10日有极端行情，收益异常高，不具有普遍参考价值，因此在回测中跳过这一天
        extreme_date = "2021-05-10"
        extreme_date = time.strptime(extreme_date, "%Y-%m-%d")
        year, month, day = extreme_date[:3]
        extreme_date = datetime.date(year, month, day)

        self.day_clearance = (self.clearance_time <= self.on_bar_time <= self.liq_time)

        if self.on_bar_date != extreme_date:

            self.bg.update_bar(bar)

            if self.day_clearance:

                if not self.buy_svt_orderids and not self.short_svt_orderids\
                    and not self.sell_svt_orderids and not self.cover_svt_orderids\
                        and not self.sell_lvt_orderids and not self.cover_lvt_orderids:
                        
                        if self.pos > 0:
                            self.sell_lvt_orderids = self.sell(self.liq_price - 5, abs(self.pos))

                        elif self.pos < 0:
                            self.cover_lvt_orderids = self.cover(self.liq_price + 5, abs(self.pos))

                else:
                    for buf_orderids in [
                        self.buy_svt_orderids,
                        self.sell_svt_orderids,
                        self.short_svt_orderids,
                        self.cover_svt_orderids,
                        self.sell_lvt_orderids,
                        self.cover_lvt_orderids
                    ]:

                        if buf_orderids:
                            for vt_orderid in buf_orderids:
                                self.cancel_order(vt_orderid)

    def on_xmin_bar(self, bar: BarData):
        """"""
        
        am = self.am
        am.update_bar(bar)
        if not am.inited:
            return

        self.boll_up, self.boll_down = am.boll(self.boll_window, self.boll_dev)

        self.ultosc = am.ultosc()
        self.buy_dis = 50 + self.dis_open
        self.short_dis = 50 - self.dis_open
        self.atr_value = am.atr(self.atr_window)

        if not self.day_clearance:

            # pos = copy.deepcopy(self.pos)

            # print(f"on_xmin_bar, self.pos:{pos} on_bar_time:{self.on_bar_time}")

            if self.pos == 0:

                self.trading_size = max(int(self.risk_level / self.atr_value), 1)
                # print(f"risk_level:{self.risk_level}, atr_value:{self.atr_value}, trading_size:{self.trading_size}")
                
                if self.trading_size > 6:
                    self.trading_size = 6

                self.intra_trade_high = bar.high_price
                self.intra_trade_low = bar.low_price

                if not self.buy_svt_orderids and not self.short_svt_orderids:
                    
                    if self.ultosc > self.buy_dis:
                        self.buy_svt_orderids = self.buy(self.boll_up, self.trading_size, True)
                        # print(f"on_xmin_bar, buy_svt:{self.buy_svt_orderids}, volume:{self.trading_size}")

                    elif self.ultosc < self.short_dis:
                        self.short_svt_orderids = self.short(self.boll_down, self.trading_size, True)
                        # print(f"on_xmin_bar, short_svt:{self.short_svt_orderids}, volume:{self.trading_size}")

                else:
                    for buf_orderids in [
                        self.buy_svt_orderids,
                        self.short_svt_orderids
                    ]:

                        if buf_orderids:
                            for vt_orderid in buf_orderids:
                                self.cancel_order(vt_orderid)
                                # print(f"on_xmin_bar, cancel {vt_orderid}")

            elif self.pos > 0:
                self.intra_trade_high = max(self.intra_trade_high, bar.high_price)
                self.intra_trade_low = bar.low_price

                self.long_stop = self.intra_trade_high - self.atr_value * self.sl_multiplier

                if not self.sell_svt_orderids:
                    self.sell_svt_orderids = self.sell(self.long_stop, abs(self.pos), True)
                    # print(f"on_xmin_bar, sell_svt:{self.sell_svt_orderids}, volume:{pos}")

                else:
                    for vt_orderid in self.sell_svt_orderids:
                        self.cancel_order(vt_orderid)
                        # print(f"on_xmin_bar, cancel {vt_orderid}")

            else:
                self.intra_trade_high = bar.high_price
                self.intra_trade_low = min(self.intra_trade_low, bar.low_price)

                self.short_stop = self.intra_trade_low + self.atr_value * self.sl_multiplier

                if not self.cover_svt_orderids:
                    self.cover_svt_orderids = self.cover(self.short_stop, abs(self.pos), True)
                    # print(f"on_xmin_bar, cover_svt:{self.cover_svt_orderids}, volume:{pos}")

                else:
                    for vt_orderid in self.cover_svt_orderids:
                        self.cancel_order(vt_orderid)
                        # print(f"on_xmin_bar, cancel {vt_orderid}")

    def on_stop_order(self, stop_order: StopOrder):
        """"""

        if stop_order.status == StopOrderStatus.WAITING:
            return

        for buf_orderids in [
            self.buy_svt_orderids,
            self.sell_svt_orderids,
            self.short_svt_orderids,
            self.cover_svt_orderids
        ]:
            
            if stop_order.stop_orderid in buf_orderids:
                buf_orderids.remove(stop_order.stop_orderid)

        if stop_order.status == StopOrderStatus.CANCELLED:

            if not self.day_clearance:

                if self.pos == 0:

                    if not self.buy_svt_orderids and not self.short_svt_orderids:

                        if self.ultosc > self.buy_dis:
                            self.buy_svt_orderids = self.buy(self.boll_up, self.trading_size, True)
                            # print(f"on_stop_order, buy_svt:{self.buy_svt_orderids}, volume:{self.trading_size}")

                        elif self.ultosc < self.short_dis:
                            self.short_svt_orderids = self.short(self.boll_down, self.trading_size, True)
                            # print(f"on_stop_order, short_svt:{self.short_svt_orderids}, volume:{self.trading_size}")

                elif self.pos > 0:
                    # pos = copy.deepcopy(self.pos)

                    if not self.sell_svt_orderids:
                        self.sell_svt_orderids = self.sell(self.long_stop, abs(self.pos), True)
                        # print(f"on_stop_order, sell_svt:{self.sell_svt_orderids}, volume:{pos}")

                else:
                    # pos = copy.deepcopy(self.pos)

                    if not self.cover_svt_orderids:  
                        self.cover_svt_orderids = self.cover(self.short_stop, abs(self.pos), True)
                        # print(f"on_stop_order, cover_svt:{self.cover_svt_orderids}, volume:{pos}")

            else:
                # pos = copy.deepcopy(self.pos)

                if not self.buy_svt_orderids and not self.short_svt_orderids\
                    and not self.sell_svt_orderids and not self.cover_svt_orderids\
                        and not self.sell_lvt_orderids and not self.cover_lvt_orderids:
                    
                        if self.pos > 0:
                            self.sell_lvt_orderids = self.sell(self.liq_price - 5, abs(self.pos))
                            # print(f"clearance time, on_stop_order, sell volume:{pos}, on_bar_time:{self.on_bar_time}")

                        elif self.pos < 0:
                            self.cover_lvt_orderids = self.cover(self.liq_price + 5, abs(self.pos))
                            # print(f"clearance time, on_stop_order, cover volume:{pos}, on_bar_time:{self.on_bar_time}")

    def on_order(self, order: OrderData):
        """"""

        if order.is_active():
            return

        for buf_orderids in [
            self.sell_lvt_orderids,
            self.cover_lvt_orderids
        ]:
            if order.orderid in buf_orderids:
                buf_orderids.remove(order.orderid)    
 
        # not ACTIVE_STATUSES = set([Status.ALLTRADED, Status.CANCELLED, Status.REJECTED])
        if order.status in [Status.CANCELLED, Status.REJECTED]:
        
            if self.day_clearance:

                # pos = copy.deepcopy(self.pos)

                if not self.buy_svt_orderids and not self.short_svt_orderids\
                    and not self.sell_svt_orderids and not self.cover_svt_orderids\
                        and not self.sell_lvt_orderids and not self.cover_lvt_orderids:
                        
                        if self.pos > 0:
                            self.sell_lvt_orderids = self.sell(self.liq_price - 5, abs(self.pos))
                            # print(f"clearance time, on_order, sell volume:{pos}, on_bar_time:{self.on_bar_time}")

                        elif self.pos < 0:
                            self.cover_lvt_orderids = self.cover(self.liq_price + 5, abs(self.pos))
                            # print(f"clearance time, on_order, cover volume:{pos}, on_bar_time:{self.on_bar_time}")

    def on_trade(self, trade: TradeData):
        """"""
        pass

        # self.trade_record_dict = {
        #     "vt_symbol": trade.vt_symbol,
        #     "orderid": trade.orderid,
        #     "tradeid": trade.tradeid,
        #     "offset": str(trade.offset),
        #     "direction": str(trade.direction),
        #     "price": trade.price,
        #     "volume": trade.volume,
        #     "datetime": str(trade.datetime),
        #     "strategy": self.strategy_name
        # }

        # self.trade_record_sheet.insert_rows(2)

        # for i in range(1, self.trade_record_sheet.max_column+1):
        #     column = get_column_letter(i)
        #     self.trade_record_sheet[column+str(2)] = list(self.trade_record_dict.values())[i-1]

        # self.trade_record_wb.save(self.path/"trade_reord_table.xlsx")

        # print("Trade Record Is Saved")


class XminBarGenerator(BarGenerator):
    def __init__(
        self,
        on_bar: Callable,
        window: int = 0,
        on_window_bar: Callable = None,
        interval: Interval = Interval.MINUTE
    ):
        super().__init__(on_bar, window, on_window_bar, interval)
    
    def update_bar(self, bar: BarData) ->None:
        """
        Update 1 minute bar into generator
        """
        # If not inited, creaate window bar object
        if not self.window_bar:
            # Generate timestamp for bar data
            if self.interval == Interval.MINUTE:
                dt = bar.datetime.replace(second=0, microsecond=0)
            else:
                dt = bar.datetime.replace(minute=0, second=0, microsecond=0)

            self.window_bar = BarData(
                symbol=bar.symbol,
                exchange=bar.exchange,
                datetime=dt,
                gateway_name=bar.gateway_name,
                open_price=bar.open_price,
                high_price=bar.high_price,
                low_price=bar.low_price
            )
        # Otherwise, update high/low price into window bar
        else:
            self.window_bar.high_price = max(
                self.window_bar.high_price, bar.high_price)
            self.window_bar.low_price = min(
                self.window_bar.low_price, bar.low_price)

        # Update close price/volume into window bar
        self.window_bar.close_price = bar.close_price
        self.window_bar.volume += int(bar.volume)
        self.window_bar.open_interest = bar.open_interest

        # Check if window bar completed
        finished = False

        if self.interval == Interval.MINUTE:
            # x-minute bar
            # if not (bar.datetime.minute + 1) % self.window:
            #     finished = True
            
            self.interval_count += 1

            if not self.interval_count % self.window:
                finished = True
                self.interval_count = 0

            elif bar.datetime.time() in [time1(10, 14), time1(11, 29), time1(14, 59), time1(22, 59)]:
                if bar.exchange in [Exchange.SHFE, Exchange.DCE, Exchange.CZCE]:
                    finished = True
                    self.interval_count = 0

        elif self.interval == Interval.HOUR:
            if self.last_bar:
                new_hour = bar.datetime.hour != self.last_bar.datetime.hour
                last_minute = bar.datetime.minute == 59
                not_first = self.window_bar.datetime != bar.datetime

                # To filter duplicate hour bar finished condition
                if (new_hour or last_minute) and not_first:
                    # 1-hour bar
                    if self.window == 1:
                        finished = True
                    # x-hour bar
                    else:
                        self.interval_count += 1

                        if not self.interval_count % self.window:
                            finished = True
                            self.interval_count = 0

        if finished:
            self.on_window_bar(self.window_bar)
            self.window_bar = None

        # Cache last bar object
        self.last_bar = bar


#%%
start1 = time.time()
engine = BacktestingEngine()
engine.set_parameters(
    vt_symbol="rb888.SHFE",
    interval="1m",
    start=datetime.datetime(2021, 1, 1),
    end=datetime.datetime(2021, 7, 26),
    rate=0.0001,
    slippage=0.2,
    size=10,
    pricetick=1,
    capital=50000,
    mode=BacktestingMode.BAR
)
engine.add_strategy(OscillatorHNBacktest, {})
#%%
start2 = time.time()
engine.load_data()
end2 = time.time()
print(f"加载数据所需时长: {(end2-start2)} Seconds")
#%%
engine.run_backtesting()
#%%
engine.calculate_result()
engine.calculate_statistics()
# 待测试的代码
end1 = time.time()
print(f"单次回测运行时长: {(end1-start1)} Seconds")
engine.show_chart()
#%%
# setting = OptimizationSetting()
# setting.set_target("end_balance")
# setting.add_parameter("bar_window_length", 1, 20, 1)
# setting.add_parameter("cci_window", 3, 10, 1)
# setting.add_parameter("fixed_size", 1, 1, 1)
# setting.add_parameter("sell_multipliaer", 0.80, 0.99, 0.01)
# setting.add_parameter("cover_multiplier", 1.01, 1.20, 0.01)
# setting.add_parameter("pricetick_multiplier", 1, 5, 1)
#%%
# engine.run_optimization(setting, output=True)